import openpyxl
import logging

class ExcelUtility:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.logger = logging.getLogger(__name__)

    def set_excel_file(self, sheet_name):
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.worksheet = self.workbook[sheet_name]

    def get_cell_data(self, row_num, col_num):
        cell_value = self.worksheet.cell(row=row_num + 1, column=col_num + 1).value  # openpyxl is 1-indexed
        return cell_value

    def set_cell_data(self, value, row_num, col_num):
        self.worksheet.cell(row=row_num + 1, column=col_num + 1).value = value
        self.workbook.save(self.file_path)
        self.logger.info(f"Set cell data at row {row_num}, col {col_num} to {value}")

    def create_excel_and_write(self, sheet_name, value):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.worksheet.cell(row=1, column=1).value = value
        self.workbook.save(self.file_path)
        self.logger.info(f"Created new excel file {self.file_path} with value {value}")
